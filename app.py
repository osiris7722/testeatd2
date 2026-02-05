from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from datetime import datetime
import sqlite3
import os
import csv
import io
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import firebase_admin
from firebase_admin import credentials, firestore, auth
import sys

app = Flask(__name__)
app.secret_key = 'sua_chave_secreta_aqui_mude_para_producao'

# Configurações
DATABASE = 'feedback.db'
ADMIN_PASSWORD = 'admin123'  # Altere esta senha!

# Admin via Firebase Auth
# - Para restringir quais contas podem entrar, defina:
#   ADMIN_EMAILS="admin1@dominio.com,admin2@dominio.com"
#   e/ou ADMIN_EMAIL_DOMAIN="dominio.com"
ADMIN_EMAILS = {
    email.strip().lower()
    for email in os.environ.get('ADMIN_EMAILS', '').split(',')
    if email.strip()
}
ADMIN_EMAIL_DOMAIN = os.environ.get('ADMIN_EMAIL_DOMAIN', '').strip().lower() or None

# Config web do Firebase (para o login no browser).
# Defina via env vars para o admin login funcionar no front-end:
# FIREBASE_API_KEY, FIREBASE_AUTH_DOMAIN, FIREBASE_PROJECT_ID, FIREBASE_APP_ID
FIREBASE_WEB_CONFIG = {
    # Defaults são os valores do teu firebaseConfig (podem ser sobrescritos por env vars)
    'apiKey': os.environ.get('FIREBASE_API_KEY', 'AIzaSyAEvUvbhv2vXj8qa1G6r9S8HSr2cFUv_bM'),
    'authDomain': os.environ.get('FIREBASE_AUTH_DOMAIN', 'studio-7634777517-713ea.firebaseapp.com'),
    'projectId': os.environ.get('FIREBASE_PROJECT_ID', 'studio-7634777517-713ea'),
    'storageBucket': os.environ.get('FIREBASE_STORAGE_BUCKET', 'studio-7634777517-713ea.firebasestorage.app'),
    'messagingSenderId': os.environ.get('FIREBASE_MESSAGING_SENDER_ID', '142898689875'),
    'appId': os.environ.get('FIREBASE_APP_ID', '1:142898689875:web:726d61b0a2590e7e4c93a6'),
    'measurementId': os.environ.get('FIREBASE_MEASUREMENT_ID', 'G-3JZQJD550E'),
}


def _is_admin_email_allowed(email: Optional[str]) -> bool:
    if not email:
        return False
    email_l = email.strip().lower()

    if ADMIN_EMAILS:
        return email_l in ADMIN_EMAILS

    if ADMIN_EMAIL_DOMAIN:
        return email_l.endswith(f"@{ADMIN_EMAIL_DOMAIN}")

    # Se não houver restrição configurada, aceita qualquer conta autenticada
    return True

# Inicializar Firebase
firebase_db = None
try:
    if not firebase_admin._apps:
        cred = credentials.Certificate('studio-7634777517-713ea-firebase-adminsdk-fbsvc-7669723ac0.json')
        firebase_admin.initialize_app(cred, {
            'databaseURL': 'https://studio-7634777517-713ea.firebaseio.com'
        })
        firebase_db = firestore.client()
        print("✓ Firebase inicializado com sucesso")
except Exception as e:
    print(f"⚠ Aviso: Firebase não está disponível: {e}")
    print("  A aplicação continuará funcionando apenas com SQLite")

def get_db():
    """Conecta ao banco de dados SQLite"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Inicializa o banco de dados"""
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            grau_satisfacao TEXT NOT NULL,
            data TEXT NOT NULL,
            hora TEXT NOT NULL,
            dia_semana TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# Inicializar banco de dados ao iniciar a aplicação
init_db()

@app.route('/')
def index():
    """Página principal com os botões de feedback"""
    return render_template('index.html')


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check simples (SQLite + Firebase init)."""
    sqlite_ok = False
    sqlite_error = None
    try:
        conn = get_db()
        conn.execute('SELECT 1').fetchone()
        conn.close()
        sqlite_ok = True
    except Exception as e:
        sqlite_error = str(e)

    firebase_ok = bool(firebase_db and firebase_admin._apps)

    return jsonify({
        'ok': sqlite_ok,
        'time': datetime.now().isoformat(),
        'python': sys.version.split(' ')[0],
        'sqlite': {
            'ok': sqlite_ok,
            'db': DATABASE,
            'error': sqlite_error,
        },
        'firebase': {
            'initialized': bool(firebase_admin._apps),
            'firestoreAvailable': bool(firebase_db),
            'ok': firebase_ok,
            'projectId': FIREBASE_WEB_CONFIG.get('projectId') or None,
        }
    })


@app.route('/api/public/summary', methods=['GET'])
def public_summary():
    """Resumo público para o ecrã principal (sem auth)."""
    try:
        hoje = datetime.now().strftime('%Y-%m-%d')
        conn = get_db()

        # Totais de hoje
        rows_hoje = conn.execute('''
            SELECT grau_satisfacao, COUNT(*) as total
            FROM feedback
            WHERE data = ?
            GROUP BY grau_satisfacao
        ''', (hoje,)).fetchall()

        # Total geral + último id
        total_geral = conn.execute('SELECT COUNT(*) as total FROM feedback').fetchone()['total']
        last_id_row = conn.execute('SELECT MAX(id) as last_id FROM feedback').fetchone()
        last_id = last_id_row['last_id'] if last_id_row else None

        conn.close()

        hoje_result = {
            'muito_satisfeito': 0,
            'satisfeito': 0,
            'insatisfeito': 0,
        }
        for r in rows_hoje:
            hoje_result[r['grau_satisfacao']] = r['total']

        return jsonify({
            'date': hoje,
            'today': hoje_result,
            'todayTotal': sum(hoje_result.values()),
            'total': total_geral,
            'lastId': last_id,
            'firebaseAvailable': bool(firebase_db),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/feedback', methods=['POST'])
def registrar_feedback():
    """Registra o feedback do usuário no SQLite e Firebase"""
    try:
        data = request.get_json()
        grau_satisfacao = data.get('grau_satisfacao')
        
        if grau_satisfacao not in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
            return jsonify({'error': 'Grau de satisfação inválido'}), 400
        
        now = datetime.now()
        data_str = now.strftime('%Y-%m-%d')
        hora_str = now.strftime('%H:%M:%S')
        timestamp_str = now.isoformat()
        
        # Dias da semana em português
        dias_semana = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 
                       'Quinta-feira', 'Sexta-feira', 'Sábado', 'Domingo']
        dia_semana = dias_semana[now.weekday()]
        
        # Preparar dados do feedback
        feedback_data = {
            'grau_satisfacao': grau_satisfacao,
            'data': data_str,
            'hora': hora_str,
            'dia_semana': dia_semana,
            'timestamp': timestamp_str
        }
        
        # Guardar no SQLite
        conn = get_db()
        cursor = conn.execute(
            'INSERT INTO feedback (grau_satisfacao, data, hora, dia_semana) VALUES (?, ?, ?, ?)',
            (grau_satisfacao, data_str, hora_str, dia_semana)
        )
        conn.commit()
        feedback_id = cursor.lastrowid
        conn.close()

        # Adicionar id ao payload (útil para Firestore e integrações)
        feedback_data['id'] = feedback_id
        
        # Guardar no Firebase (Firestore)
        if firebase_db:
            try:
                firebase_db.collection('feedback').document(f'feedback_{feedback_id}').set(feedback_data)
                print(f"✓ Feedback {feedback_id} sincronizado com Firebase")
            except Exception as firebase_error:
                # Log do erro mas não falha a resposta
                print(f"⚠ Aviso: Erro ao guardar no Firebase: {firebase_error}")
        
        return jsonify({
            'success': True,
            'message': 'Obrigado pelo seu feedback!',
            'id': feedback_id
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin_rocha', methods=['GET', 'POST'])
def admin_login():
    """Página de login do admin"""
    next_url = request.args.get('next')
    # Login do admin agora é via Firebase Authentication (no browser) + verificação do token no backend.
    if request.method == 'POST':
        return render_template(
            'admin_login.html',
            error='O login agora é feito via Firebase Authentication.',
            firebase_web_config=FIREBASE_WEB_CONFIG,
            logout_requested=bool(request.args.get('logout')),
            admin_email_domain=ADMIN_EMAIL_DOMAIN,
            admin_emails=sorted(list(ADMIN_EMAILS)),
            next_url=next_url,
        )
    
    # Se já estiver logado, redireciona para o dashboard
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_dashboard'))
    
    return render_template(
        'admin_login.html',
        firebase_web_config=FIREBASE_WEB_CONFIG,
        logout_requested=bool(request.args.get('logout')),
        admin_email_domain=ADMIN_EMAIL_DOMAIN,
        admin_emails=sorted(list(ADMIN_EMAILS)),
        next_url=next_url,
    )

@app.route('/admin_rocha/dashboard')
def admin_dashboard():
    """Dashboard administrativo"""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    return render_template('admin_dashboard.html', admin_email=session.get('admin_email'), firebase_web_config=FIREBASE_WEB_CONFIG)


@app.route('/admin_rocha/tv')
def admin_tv():
    """Modo TV (kiosk) do dashboard."""
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login', next='/admin_rocha/tv'))

    return render_template(
        'admin_tv.html',
        admin_email=session.get('admin_email'),
    )

@app.route('/admin_rocha/logout')
def admin_logout():
    """Logout do admin"""
    session.pop('admin_logged_in', None)
    session.pop('admin_uid', None)
    session.pop('admin_email', None)
    # Redireciona com flag para o front-end fazer signOut() do Firebase, se estiver autenticado.
    return redirect(url_for('admin_login', logout=1))


@app.route('/api/admin/login/firebase', methods=['POST'])
def admin_login_firebase():
    """Cria sessão admin a partir de um Firebase ID token (verificado no backend)."""
    try:
        body = request.get_json(silent=True) or {}
        id_token = body.get('idToken')

        if not id_token:
            return jsonify({'error': 'idToken ausente'}), 400

        if not firebase_admin._apps:
            return jsonify({'error': 'Firebase não inicializado'}), 503

        decoded = auth.verify_id_token(id_token)
        uid = decoded.get('uid')
        email = decoded.get('email')

        if not email and uid:
            user = auth.get_user(uid)
            email = user.email

        if not _is_admin_email_allowed(email):
            return jsonify({'error': 'Conta não autorizada para admin'}), 403

        session['admin_logged_in'] = True
        session['admin_uid'] = uid
        session['admin_email'] = email

        return jsonify({'success': True, 'email': email})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/me', methods=['GET'])
def admin_me():
    """Retorna info da sessão admin atual."""
    if not session.get('admin_logged_in'):
        return jsonify({'loggedIn': False}), 200
    return jsonify({'loggedIn': True, 'email': session.get('admin_email')}), 200

@app.route('/api/admin/stats')
def get_stats():
    """Retorna estatísticas gerais"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        conn = get_db()
        
        # Total por tipo de satisfação
        stats = conn.execute('''
            SELECT grau_satisfacao, COUNT(*) as total
            FROM feedback
            GROUP BY grau_satisfacao
        ''').fetchall()
        
        # Total geral
        total_geral = conn.execute('SELECT COUNT(*) as total FROM feedback').fetchone()['total']
        
        conn.close()
        
        # Calcular percentagens
        resultado = {
            'muito_satisfeito': 0,
            'satisfeito': 0,
            'insatisfeito': 0,
            'total': total_geral
        }
        
        percentagens = {
            'muito_satisfeito': 0,
            'satisfeito': 0,
            'insatisfeito': 0
        }
        
        for row in stats:
            grau = row['grau_satisfacao']
            total = row['total']
            resultado[grau] = total
            if total_geral > 0:
                percentagens[grau] = round((total / total_geral) * 100, 2)
        
        resultado['percentagens'] = percentagens
        
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/stats/daily')
def get_daily_stats():
    """Retorna estatísticas por dia"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        data_filtro = request.args.get('data')
        
        conn = get_db()
        
        if data_filtro:
            stats = conn.execute('''
                SELECT grau_satisfacao, COUNT(*) as total
                FROM feedback
                WHERE data = ?
                GROUP BY grau_satisfacao
            ''', (data_filtro,)).fetchall()
        else:
            # Retorna o dia atual
            hoje = datetime.now().strftime('%Y-%m-%d')
            stats = conn.execute('''
                SELECT grau_satisfacao, COUNT(*) as total
                FROM feedback
                WHERE data = ?
                GROUP BY grau_satisfacao
            ''', (hoje,)).fetchall()
        
        conn.close()
        
        resultado = {
            'muito_satisfeito': 0,
            'satisfeito': 0,
            'insatisfeito': 0
        }
        
        for row in stats:
            resultado[row['grau_satisfacao']] = row['total']
        
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/stats/comparison')
def get_comparison_stats():
    """Retorna comparação entre dois períodos"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        data1_inicio = request.args.get('data1_inicio')
        data1_fim = request.args.get('data1_fim')
        data2_inicio = request.args.get('data2_inicio')
        data2_fim = request.args.get('data2_fim')
        
        if not all([data1_inicio, data1_fim, data2_inicio, data2_fim]):
            return jsonify({'error': 'Datas inválidas'}), 400
        
        conn = get_db()
        
        # Período 1
        stats1 = conn.execute('''
            SELECT grau_satisfacao, COUNT(*) as total
            FROM feedback
            WHERE data BETWEEN ? AND ?
            GROUP BY grau_satisfacao
        ''', (data1_inicio, data1_fim)).fetchall()
        
        # Período 2
        stats2 = conn.execute('''
            SELECT grau_satisfacao, COUNT(*) as total
            FROM feedback
            WHERE data BETWEEN ? AND ?
            GROUP BY grau_satisfacao
        ''', (data2_inicio, data2_fim)).fetchall()
        
        conn.close()
        
        # Formatar resultado
        resultado = {
            'periodo1': {
                'muito_satisfeito': 0,
                'satisfeito': 0,
                'insatisfeito': 0,
                'total': 0
            },
            'periodo2': {
                'muito_satisfeito': 0,
                'satisfeito': 0,
                'insatisfeito': 0,
                'total': 0
            }
        }
        
        for row in stats1:
            resultado['periodo1'][row['grau_satisfacao']] = row['total']
            resultado['periodo1']['total'] += row['total']
        
        for row in stats2:
            resultado['periodo2'][row['grau_satisfacao']] = row['total']
            resultado['periodo2']['total'] += row['total']
        
        # Calcular variações percentuais
        resultado['variacao'] = {}
        for key in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
            val1 = resultado['periodo1'][key]
            val2 = resultado['periodo2'][key]
            
            if val1 == 0:
                variacao = 100 if val2 > 0 else 0
            else:
                variacao = round(((val2 - val1) / val1) * 100, 2)
            
            resultado['variacao'][key] = variacao
        
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/historico')
def get_historico():
    """Retorna o histórico de feedbacks"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        offset = (page - 1) * per_page

        # Filtros opcionais (não quebram o comportamento atual)
        q = (request.args.get('q') or '').strip()
        grau = (request.args.get('grau') or '').strip()
        data_inicio = (request.args.get('data_inicio') or '').strip()
        data_fim = (request.args.get('data_fim') or '').strip()
        
        conn = get_db()
        
        where = []
        params = []

        if grau in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
            where.append('grau_satisfacao = ?')
            params.append(grau)

        if data_inicio and data_fim:
            where.append('data BETWEEN ? AND ?')
            params.extend([data_inicio, data_fim])
        elif data_inicio:
            where.append('data >= ?')
            params.append(data_inicio)
        elif data_fim:
            where.append('data <= ?')
            params.append(data_fim)

        if q.isdigit():
            where.append('id = ?')
            params.append(int(q))

        where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''

        # Total de registros (com filtros)
        total = conn.execute(f'SELECT COUNT(*) as total FROM feedback{where_sql}', tuple(params)).fetchone()['total']

        # Registros paginados (com filtros)
        registros = conn.execute(f'''
            SELECT id, grau_satisfacao, data, hora, dia_semana
            FROM feedback
            {where_sql}
            ORDER BY data DESC, hora DESC
            LIMIT ? OFFSET ?
        ''', tuple(params + [per_page, offset])).fetchall()
        
        conn.close()
        
        resultado = {
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page,
            'registros': [dict(row) for row in registros]
        }
        
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/export/csv')
@app.route('/api/admin/export/xlsx')
def export_csv():
    """Exporta dados em formato Excel"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        conn = get_db()
        
        if data_inicio and data_fim:
            registros = conn.execute('''
                SELECT id, grau_satisfacao, data, hora, dia_semana
                FROM feedback
                WHERE data BETWEEN ? AND ?
                ORDER BY data, hora
            ''', (data_inicio, data_fim)).fetchall()
        else:
            registros = conn.execute('''
                SELECT id, grau_satisfacao, data, hora, dia_semana
                FROM feedback
                ORDER BY data, hora
            ''').fetchall()
        
        conn.close()
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback"
        
        # Cabeçalho com formatação
        header = ['ID', 'Grau de Satisfação', 'Data', 'Hora', 'Dia da Semana']
        ws.append(header)
        
        # Formatar cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Adicionar dados
        grau_map = {
            'muito_satisfeito': 'Muito Satisfeito',
            'satisfeito': 'Satisfeito',
            'insatisfeito': 'Insatisfeito'
        }
        
        for row in registros:
            ws.append([
                row['id'],
                grau_map.get(row['grau_satisfacao'], row['grau_satisfacao']),
                row['data'],
                row['hora'],
                row['dia_semana']
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        
        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'feedback_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/export/csv-plain')
def export_csv_plain():
    """Exporta dados em CSV (texto). Mantém o /export/csv antigo como Excel."""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401

    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')

        conn = get_db()
        if data_inicio and data_fim:
            registros = conn.execute('''
                SELECT id, grau_satisfacao, data, hora, dia_semana
                FROM feedback
                WHERE data BETWEEN ? AND ?
                ORDER BY data, hora
            ''', (data_inicio, data_fim)).fetchall()
        else:
            registros = conn.execute('''
                SELECT id, grau_satisfacao, data, hora, dia_semana
                FROM feedback
                ORDER BY data, hora
            ''').fetchall()
        conn.close()

        grau_map = {
            'muito_satisfeito': 'Muito Satisfeito',
            'satisfeito': 'Satisfeito',
            'insatisfeito': 'Insatisfeito'
        }

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['id', 'grau_satisfacao', 'data', 'hora', 'dia_semana'])
        for row in registros:
            writer.writerow([
                row['id'],
                grau_map.get(row['grau_satisfacao'], row['grau_satisfacao']),
                row['data'],
                row['hora'],
                row['dia_semana'],
            ])

        csv_bytes = output.getvalue().encode('utf-8')
        return send_file(
            io.BytesIO(csv_bytes),
            mimetype='text/csv; charset=utf-8',
            as_attachment=True,
            download_name=f'feedback_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/system')
def admin_system():
    """Info de sistema para o dashboard (requer admin)."""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401

    try:
        conn = get_db()
        total = conn.execute('SELECT COUNT(*) as total FROM feedback').fetchone()['total']
        last_id_row = conn.execute('SELECT MAX(id) as last_id FROM feedback').fetchone()
        last_id = last_id_row['last_id'] if last_id_row else None
        conn.close()

        db_size = None
        try:
            if os.path.exists(DATABASE):
                db_size = os.path.getsize(DATABASE)
        except Exception:
            db_size = None

        return jsonify({
            'time': datetime.now().isoformat(),
            'python': sys.version.split(' ')[0],
            'total': total,
            'lastId': last_id,
            'db': {
                'path': DATABASE,
                'sizeBytes': db_size,
            },
            'firebase': {
                'initialized': bool(firebase_admin._apps),
                'firestoreAvailable': bool(firebase_db),
                'projectId': FIREBASE_WEB_CONFIG.get('projectId') or None,
            },
            'admin': {
                'email': session.get('admin_email'),
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/export/txt')
def export_txt():
    """Exporta dados em formato TXT"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        conn = get_db()
        
        if data_inicio and data_fim:
            registros = conn.execute('''
                SELECT id, grau_satisfacao, data, hora, dia_semana
                FROM feedback
                WHERE data BETWEEN ? AND ?
                ORDER BY data, hora
            ''', (data_inicio, data_fim)).fetchall()
        else:
            registros = conn.execute('''
                SELECT id, grau_satisfacao, data, hora, dia_semana
                FROM feedback
                ORDER BY data, hora
            ''').fetchall()
        
        conn.close()
        
        # Criar TXT em memória
        output = io.StringIO()
        output.write('=' * 80 + '\n')
        output.write('RELATÓRIO DE FEEDBACK DE SATISFAÇÃO\n')
        output.write('=' * 80 + '\n\n')
        
        grau_map = {
            'muito_satisfeito': 'Muito Satisfeito',
            'satisfeito': 'Satisfeito',
            'insatisfeito': 'Insatisfeito'
        }
        
        for row in registros:
            output.write(f"ID: {row['id']}\n")
            output.write(f"Grau de Satisfação: {grau_map.get(row['grau_satisfacao'], row['grau_satisfacao'])}\n")
            output.write(f"Data: {row['data']}\n")
            output.write(f"Hora: {row['hora']}\n")
            output.write(f"Dia da Semana: {row['dia_semana']}\n")
            output.write('-' * 80 + '\n\n')
        
        output.write(f"\nTotal de registros: {len(registros)}\n")
        output.write(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        
        # Preparar arquivo para download
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/plain',
            as_attachment=True,
            download_name=f'feedback_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/dates')
def get_available_dates():
    """Retorna as datas disponíveis para filtragem"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Não autorizado'}), 401
    
    try:
        conn = get_db()
        dates = conn.execute('''
            SELECT DISTINCT data
            FROM feedback
            ORDER BY data DESC
        ''').fetchall()
        conn.close()
        
        return jsonify([row['data'] for row in dates])
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', '8000'))
    app.run(host='127.0.0.1', port=port, debug=True)
